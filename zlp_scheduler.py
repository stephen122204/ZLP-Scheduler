#!/usr/bin/env python3
"""
Author - Stephen Abkin, Class of 2027, Cohort J
Date - 05/04/2025

zlp_scheduler.py — ZLP 100-minute window finder with minimum conflict reporting.
Upgrades December 19, 2025:
  - Bundled lecture+lab sections (inseparable)
  - Export top 10 and a heatmap to zlp_results.xlsx
  - Print top 10 in terminal
  - Remove obsolete manual input 
"""

from __future__ import annotations
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass
from copy import deepcopy

# ─────────────────────────────────────────────────────────────
# Basic constants that describe the week and the time grid
# ─────────────────────────────────────────────────────────────
DAY_LETTERS = "MTWRF"
DAY_NAMES   = {"M":"Monday","T":"Tuesday","W":"Wednesday","R":"Thursday","F":"Friday"}

GRID_START, GRID_END = 8*60, 16*60 + 10
BLOCK_LEN = 100
STEP_MIN = 5

CRS_RE  = re.compile(r"^[A-Z]{4}$")
NUM_RE  = re.compile(r"^\d{3}[Ll]?$")
TIME_RE = re.compile(r"^(2[0-3]|1\d|0\d):([0-5]\d)$")
DAYS_RE = re.compile(r"^[MTWRF]+$", re.I)

DEFAULT_SHEET = "sections.xlsx"
OUT_XLSX = "zlp_results.xlsx"

# ─────────────────────────────────────────────────────────────
# Data model: each course has multiple "options" (sections).
# Each option can include multiple meetings (lecture + lab).
# ─────────────────────────────────────────────────────────────
@dataclass(frozen=True)
class Meeting:
    days: str
    start: int
    dur: int
    label: str  # e.g., "MEEN 221" or "MEEN 221 (Lab)"

@dataclass(frozen=True)
class Option:
    course: str
    meetings: Tuple[Meeting, ...]  # inseparable bundle

# ─────────────────────────────────────────────────────────────
# Small time helpers
# ─────────────────────────────────────────────────────────────
def to_minutes(hhmm: str) -> int:
    return int(hhmm[:2])*60 + int(hhmm[3:])

def to_hhmm(m: int) -> str:
    return f"{m//60:02d}:{m%60:02d}"

def overlaps(a:Tuple[int,int], b:Tuple[int,int]) -> bool:
    return max(a[0],b[0]) < min(a[1],b[1])

def merge(intvs: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    merged: List[Tuple[int, int]] = []
    for s, e in sorted(intvs):
        if not merged or s > merged[-1][1]:
            merged.append((s, e))
        else:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))
    return merged

def free_and_min_conflict(day:List[Tuple[int,int]]):
    free, best, mincnt = [], [], float('inf')
    for st in range(GRID_START, GRID_END+1, STEP_MIN):
        blk=(st, st+BLOCK_LEN)
        cnt=sum(1 for iv in day if overlaps(blk,iv))
        if cnt==0:
            free.append(blk)
        if cnt<mincnt:
            mincnt,best=cnt,[st]
        elif cnt==mincnt:
            best.append(st)
    return free,best,mincnt

def format_start_span(starts: List[int]) -> str:
    if not starts:
        return "(none)"
    starts = sorted(starts)
    first = starts[0]
    last = starts[-1]
    if len(starts) == 1:
        return to_hhmm(first)
    return f"{to_hhmm(first)}–{to_hhmm(last)} (every 5 min)"

# ─────────────────────────────────────────────────────────────
# Spreadsheet load
# Supports BOTH formats:
#   Old columns: Subject, Number, Days, Start, Duration
#   New columns: + Lab, Lab_Days, Lab_Start, Lab_Duration
# Lab columns are only read/validated if Lab is truthy (Y/YES/TRUE/1).
# ─────────────────────────────────────────────────────────────
def rows_from_file(path: str) -> List[dict]:
    import pandas as pd

    def norm_cell(x) -> str:
        return "" if pd.isna(x) else str(x).strip()

    p = Path(path)
    if p.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(p, engine="openpyxl")
    elif p.suffix.lower() == ".csv":
        df = pd.read_csv(p)
    else:
        raise ValueError("file must be .xlsx, .xls, or .csv")

    need = ["Subject", "Number", "Days", "Start", "Duration"]
    miss = [c for c in need if c not in df.columns]
    if miss:
        raise ValueError(f"missing columns: {', '.join(miss)}")

    has_lab_cols = all(c in df.columns for c in ["Lab", "Lab_Days", "Lab_Start", "Lab_Duration"])

    out: List[dict] = []
    for _, r in df.iterrows():
        subj = norm_cell(r["Subject"]).upper()
        num = norm_cell(r["Number"])
        days = norm_cell(r["Days"]).upper()
        start = norm_cell(r["Start"])

        dur_raw = r["Duration"]
        if pd.isna(dur_raw) or str(dur_raw).strip() == "":
            duration = ""
        else:
            duration = str(int(float(dur_raw))).strip()

        row = {
            "Subject": subj,
            "Number": num,
            "Days": days,
            "Start": start,
            "Duration": duration,
            "Lab": "",
            "Lab_Days": "",
            "Lab_Start": "",
            "Lab_Duration": "",
        }

        if has_lab_cols:
            lab_flag = norm_cell(r["Lab"]).upper()
            row["Lab"] = lab_flag

            # Only parse lab details if Lab is truthy
            if lab_flag in {"Y", "YES", "TRUE", "1"}:
                row["Lab_Days"] = norm_cell(r["Lab_Days"]).upper()
                row["Lab_Start"] = norm_cell(r["Lab_Start"])

                lab_dur_raw = r["Lab_Duration"]
                if pd.isna(lab_dur_raw) or str(lab_dur_raw).strip() == "":
                    row["Lab_Duration"] = ""
                else:
                    row["Lab_Duration"] = str(int(float(lab_dur_raw))).strip()

        out.append(row)

    return out


def validate_meeting(days: str, start: str, dur: str) -> Tuple[str,int,int]:
    days = days.upper()
    if not DAYS_RE.fullmatch(days):
        raise ValueError("days must be combo of MTWRF")
    if not TIME_RE.fullmatch(start):
        raise ValueError("start must be HH:MM 24-hour")
    if not dur.isdigit() or int(dur) <= 0:
        raise ValueError("duration must be positive int")
    st = to_minutes(start)
    return days, st, int(dur)

def add_row_as_option(row: dict, sections: Dict[str, List[Option]]) -> bool:
    """
    Adds ONE spreadsheet row as ONE "Option" for a course.
    If Lab=Y, bundles lab meeting into the same Option.
    """
    subj = row["Subject"]
    num  = row["Number"]
    code = f"{subj} {num}"

    try:
        if not CRS_RE.fullmatch(subj) or not NUM_RE.fullmatch(num):
            raise ValueError("course code malformed (e.g. MEEN 221)")

        lec_days, lec_st, lec_dur = validate_meeting(row["Days"], row["Start"], row["Duration"])
        meetings: List[Meeting] = [Meeting(lec_days, lec_st, lec_dur, code)]

        lab_flag = str(row.get("Lab","")).strip().upper()
        if lab_flag in {"Y","YES","TRUE","1"}:
            lab_days = str(row.get("Lab_Days","")).strip().upper()
            lab_start = str(row.get("Lab_Start","")).strip()
            lab_dur = str(row.get("Lab_Duration","")).strip()
            if not (lab_days and lab_start and lab_dur):
                raise ValueError("Lab=Y but lab fields are missing (Lab_Days/Lab_Start/Lab_Duration)")
            ld, ls, ldur = validate_meeting(lab_days, lab_start, lab_dur)
            meetings.append(Meeting(ld, ls, ldur, f"{code} (Lab)"))

        sections.setdefault(code, []).append(Option(course=code, meetings=tuple(meetings)))
        return True

    except ValueError as err:
        print(f"Error in {code}: {err}")
        return False

# ─────────────────────────────────────────────────────────────
# Scheduling scoring + optimistic conflict check (now option-aware)
# ─────────────────────────────────────────────────────────────
def windows_after_add(
    cand: Option,
    grid:Dict[str,List[Tuple[int,int]]],
    grid_raw:Dict[str,List[Tuple[int,int,str]]]
) -> Tuple[int,int]:
    tmp = deepcopy(grid)
    tmp_raw = deepcopy(grid_raw)

    # add all meetings in this option (lecture + lab)
    for mtg in cand.meetings:
        iv = (mtg.start, mtg.start + mtg.dur)
        for d in mtg.days:
            tmp[d].append(iv)
            tmp[d] = merge(tmp[d])
            tmp_raw[d].append((iv[0], iv[1], cand.course))

    total_free = 0
    total_min_conf = 0
    for d in DAY_LETTERS:
        free, _, _ = free_and_min_conflict(tmp[d])
        total_free += len(free)

        # label-aware min conflicts sweep
        mincnt = min(
            sum(1 for s,e,_ in tmp_raw[d] if overlaps((s,e),(st2,st2+BLOCK_LEN)))
            for st2 in range(GRID_START, GRID_END+1, STEP_MIN)
        )
        total_min_conf += mincnt

    return total_free, total_min_conf

def option_has_nonoverlap_on_day(opt: Option, day_letter: str, blk: Tuple[int,int]) -> bool:
    """
    True if this option can exist on this day without overlapping the block.
    If the option has no meetings on that day, it's trivially OK.
    If it has meetings on that day, ALL of them must not overlap the block.
    """
    day_meetings = [m for m in opt.meetings if day_letter in m.days]
    if not day_meetings:
        return True
    for m in day_meetings:
        if overlaps((m.start, m.start + m.dur), blk):
            return False
    return True

def adjusted_conflicts_for_start(
    st: int,
    day_labeled: List[Tuple[int,int,str]],
    sections_all: Dict[str, List[Option]],
    day_letter: str
) -> List[Tuple[str,int,int]]:
    """
    Optimistic check:
    - Find all courses that overlap the block on this day (raw hits).
    - If that course has some OTHER option that avoids overlapping this day/block
      (including bundled lab), then treat it as movable and don't count it.
    """
    blk = (st, st + BLOCK_LEN)

    raw_hits = [(code, s, e) for (s, e, code) in day_labeled if overlaps((s, e), blk)]
    remaining: List[Tuple[str,int,int]] = []

    for code, s, e in raw_hits:
        movable = False
        if code in sections_all:
            for opt in sections_all[code]:
                # if some option for this course avoids overlap on this day, it's movable
                if option_has_nonoverlap_on_day(opt, day_letter, blk):
                    movable = True
                    break
        if not movable:
            remaining.append((code, s, e))

    return remaining

def option_overlaps_block(opt: Option, day: str, blk: Tuple[int, int]) -> bool:
    """
    True if this option has any meeting on 'day' that overlaps blk.
    Labs are already bundled into opt.meetings, so this respects inseparability.
    """
    for m in opt.meetings:
        if day in m.days:
            if overlaps((m.start, m.start + m.dur), blk):
                return True
    return False

def best_conflicts_for_block(
    day: str,
    st: int,
    sections_all: Dict[str, List[Option]]
) -> Tuple[int, List[str]]:
    """
    For a fixed meeting block (day, st..st+BLOCK_LEN), compute the minimum possible
    number of conflicting courses assuming we can re-pick one option per course
    specifically for this start time.

    A course is an unavoidable conflict if every option overlaps the meeting block.
    Returns (score, sorted_conflict_courses).
    """
    blk = (st, st + BLOCK_LEN)
    conflicts: List[str] = []

    for course, opts in sections_all.items():
        can_avoid = any(not option_overlaps_block(opt, day, blk) for opt in opts)
        if not can_avoid:
            conflicts.append(course)

    conflicts = sorted(set(conflicts))
    return len(conflicts), conflicts

# blocked courses are those that lose at least one option but are not full conflicts (LOGIC)
def best_conflicts_and_blocked_for_block(
    day: str,
    st: int,
    sections_all: Dict[str, List[Option]]
) -> Tuple[int, List[str], List[str]]:
    """
    For a fixed meeting block, compute:
      - score: minimum unavoidable conflicts (courses where every option overlaps)
      - conflicts: those unavoidable conflict courses (subject+number)
      - blocked: courses that remain doable but lose at least one option due to overlap

    Blocked means: at least one option overlaps AND at least one option does not overlap.
    """
    blk = (st, st + BLOCK_LEN)

    conflicts: List[str] = []
    blocked: List[str] = []

    for course, opts in sections_all.items():
        any_overlap = any(option_overlaps_block(opt, day, blk) for opt in opts)
        any_clear   = any(not option_overlaps_block(opt, day, blk) for opt in opts)

        if not any_clear:
            # Every option overlaps -> unavoidable conflict
            conflicts.append(course)
        elif any_overlap:
            # Still possible, but choices are restricted
            blocked.append(course)

    conflicts = sorted(set(conflicts))
    blocked = sorted(set(blocked))
    return len(conflicts), conflicts, blocked

# ─────────────────────────────────────────────────────────────
# Reporting: top 10 + heatmap spreadsheet
# ─────────────────────────────────────────────────────────────
def generate_reports(sections_all: Dict[str, List[Option]]) -> None:
    """
    Meeting-time-centric reporting.

    For each candidate (day, start), we re-pick course options independently to
    minimize conflicts with the 100-minute meeting block.

    Outputs:
      - Terminal: all contiguous ranges with score <= 2; if fewer than 10, fills with next best ranges
      - Excel: one sheet with Heatmap (time rows, day cols) + table starting at column H
      - Table includes unavoidable conflicting courses (subject + number) for that block
      - Table includes blocked courses and a blocked count column
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    starts = list(range(GRID_START, GRID_END + 1, STEP_MIN))
    time_rows = [to_hhmm(s) for s in starts]

    heat_rows = {d: [] for d in DAY_LETTERS}

    conflict_map: Dict[Tuple[str, int], List[str]] = {}
    blocked_map: Dict[Tuple[str, int], List[str]] = {}
    score_map: Dict[Tuple[str, int], int] = {}

    for d in DAY_LETTERS:
        for s in starts:
            score, conflicts, blocked = best_conflicts_and_blocked_for_block(d, s, sections_all)
            conflict_map[(d, s)] = conflicts
            blocked_map[(d, s)] = blocked
            score_map[(d, s)] = score
            heat_rows[d].append(score)

    df_heat = pd.DataFrame(heat_rows, index=time_rows)
    df_heat.index.name = "StartTime"

    def group_contiguous_starts(
        day: str,
        score: int,
        st_list: List[int]
    ) -> List[Tuple[int, int, str, int, int, int, str, int, str]]:
        """
        Tuple:
          (score, -count, day, first_start, last_start, count, conflicts_str, blocked_count, blocked_str)
        """
        st_list = sorted(st_list)
        out: List[Tuple[int, int, str, int, int, int, str, int, str]] = []
        if not st_list:
            return out

        first = prev = st_list[0]
        count = 1

        for x in st_list[1:]:
            if x == prev + STEP_MIN:
                prev = x
                count += 1
            else:
                conflicts_str = ", ".join(conflict_map[(day, first)])
                blocked_list = blocked_map[(day, first)]
                blocked_count = len(blocked_list)
                blocked_str = ", ".join(blocked_list)

                out.append((score, -count, day, first, prev, count, conflicts_str, blocked_count, blocked_str))

                first = prev = x
                count = 1

        conflicts_str = ", ".join(conflict_map[(day, first)])
        blocked_list = blocked_map[(day, first)]
        blocked_count = len(blocked_list)
        blocked_str = ", ".join(blocked_list)

        out.append((score, -count, day, first, prev, count, conflicts_str, blocked_count, blocked_str))
        return out

    # Build ranges across all days and all scores
    ranges: List[Tuple[int, int, str, int, int, int, str, int, str]] = []
    for d in DAY_LETTERS:
        by_score: Dict[int, List[int]] = {}
        for s in starts:
            sc = score_map[(d, s)]
            by_score.setdefault(sc, []).append(s)

        for sc, st_list in by_score.items():
            ranges.extend(group_contiguous_starts(d, sc, st_list))

    # Sort: score -> #blocked -> earlier start -> earlier day:
    ranges.sort(key=lambda t: (t[0], t[7], t[3], DAY_LETTERS.index(t[2])))

    # Selection rule: include all score <= 2; if fewer than 10, fill with next best
    selected = [t for t in ranges if t[0] <= 2]
    if len(selected) < 10:
        need = 10 - len(selected)
        extras = [t for t in ranges if t[0] > 2]
        selected.extend(extras[:need])

    # Terminal print
    print("\nMeeting start-time ranges (contiguous, per-start optimized):")
    print("Includes all ranges with score <= 2; if fewer than 10, filled with next best ranges.\n")

    for i, (score, _, d, first, last, count, conflicts_str, blocked_count, blocked_str) in enumerate(selected, 1):
        if first == last:
            start_range = to_hhmm(first)
        else:
            start_range = f"{to_hhmm(first)}–{to_hhmm(last)}"

        print(f"{i:2d}. {DAY_NAMES[d]:9s}   start: {start_range}    score={score}")

    # Write one sheet
    out_path = Path(__file__).with_name(OUT_XLSX)
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df_heat.to_excel(w, sheet_name="ScheduleData")

    wb = load_workbook(out_path)
    ws = wb["ScheduleData"]

    center = Alignment(horizontal="center", vertical="center")

    # Heatmap formatting
    ws.freeze_panes = "B2"
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = center

    for c in range(2, 2 + len(DAY_LETTERS)):
        hdr = ws.cell(row=1, column=c)
        day_letter = hdr.value
        if day_letter in DAY_NAMES:
            hdr.value = f"{day_letter} ({DAY_NAMES[day_letter]})"
        hdr.font = Font(bold=True)
        hdr.alignment = center

    ws.column_dimensions["A"].width = 10
    for r in range(2, 2 + len(time_rows)):
        ws.cell(row=r, column=1).alignment = center

    for c in range(2, 2 + len(DAY_LETTERS)):
        ws.column_dimensions[get_column_letter(c)].width = 14
        for r in range(2, 2 + len(time_rows)):
            ws.cell(row=r, column=c).alignment = center

    # Conditional formatting (B2..F last)
    start_cell = ws.cell(row=2, column=2).coordinate
    end_cell = ws.cell(row=1 + len(time_rows), column=1 + len(DAY_LETTERS)).coordinate
    rng = f"{start_cell}:{end_cell}"

    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="F8696B"
    )
    ws.conditional_formatting.add(rng, rule)

    # Top table at column H
    top_col = 8  # H
    top_row = 1

    # Clear prior table area
    max_clear_rows = 300
    max_clear_cols = 20
    for rr in range(top_row, top_row + max_clear_rows):
        for cc in range(top_col, top_col + max_clear_cols):
            ws.cell(row=rr, column=cc).value = None

    headers = [
        "Rank", "Day", "Start range", "End range",
        "Range length (starts)", "Score (conflicts)",
        "Conflicting courses", "# Blocked Courses", "Blocked courses"
    ]

    for j, h in enumerate(headers):
        cell = ws.cell(row=top_row, column=top_col + j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = center

    for i, (score, _, d, first, last, count, conflicts_str, blocked_count, blocked_str) in enumerate(selected, 1):
        if first == last:
            start_range = to_hhmm(first)
            end_range = to_hhmm(first + BLOCK_LEN)
        else:
            start_range = f"{to_hhmm(first)}–{to_hhmm(last)}"
            end_range = f"{to_hhmm(first + BLOCK_LEN)}–{to_hhmm(last + BLOCK_LEN)}"

        row_vals = [
            i,
            DAY_NAMES[d],
            start_range,
            end_range,
            count,
            score,
            conflicts_str,
            blocked_count,
            blocked_str
        ]

        rr = top_row + i
        for j, val in enumerate(row_vals):
            cell = ws.cell(row=rr, column=top_col + j, value=val)
            cell.alignment = center

    widths = [6, 12, 16, 16, 20, 16, 35, 16, 120]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(top_col + j)].width = w

    wb.save(out_path)
    print(f"\nExcel report written to: {out_path}")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main() -> None:
    sections: Dict[str, List[Option]] = {}

    # 1) Auto-load sections.xlsx in same folder, else ask for a path and require it
    def_sheet = Path(__file__).with_name(DEFAULT_SHEET)
    path_to_use: Optional[Path] = None

    if def_sheet.exists():
        path_to_use = def_sheet
    else:
        p = input("Spreadsheet path (.xlsx/.csv): ").strip()
        if p:
            path_to_use = Path(p)

    if not path_to_use or not path_to_use.exists():
        print("\nNo spreadsheet found/provided; nothing to compute.")
        return

    try:
        rows = rows_from_file(str(path_to_use))
        ok = 0
        for row in rows:
            if add_row_as_option(row, sections):
                ok += 1
        if ok == 0:
            print("\nNo valid rows loaded; nothing to compute.")
            return
        print(f"[spreadsheet] loaded {ok} section rows successfully from {path_to_use.name}")
    except Exception as e:
        print(f"[file-load error] {e}")
        return

    # keep all options for optimistic check
    sections_all = sections

    # 2) Generate reports
    generate_reports(sections_all)

if __name__=="__main__":
    main()